from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook


def read_sequence_xlsx(path: str) -> List[Tuple[int, int]]:
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Sequence file not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sequence file is empty.")

    header = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    required = {"chapter_number", "verse_number"}
    if not required.issubset(set(header)):
        raise ValueError("Sequence file must include chapter_number and verse_number headers.")

    chapter_idx = header.index("chapter_number")
    verse_idx = header.index("verse_number")

    sequence: List[Tuple[int, int]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        chapter_value = row[chapter_idx] if chapter_idx < len(row) else None
        verse_value = row[verse_idx] if verse_idx < len(row) else None
        if chapter_value is None or verse_value is None:
            raise ValueError(f"Blank row detected at Excel row {row_index}.")
        try:
            chapter_number = int(chapter_value)
            verse_number = int(verse_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid numbers at Excel row {row_index}.") from exc
        sequence.append((chapter_number, verse_number))

    if not sequence:
        raise ValueError("Sequence file contains no verse rows.")

    return sequence
